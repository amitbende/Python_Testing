from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time
from selenium.webdriver.common.by import By

# Open Browser
chrome_options = Options()
chrome_options.add_argument("--remote-allow-origins=*")
chrome_options.add_argument("ignore-certificate-errors")
driver = webdriver.Chrome(options=chrome_options)

# Window Maximize
driver.maximize_window()

# Open the website
driver.get("http://adactinhotelapp.com/HotelAppBuild2/index.php")

# Wait
time.sleep(2)

# Open Excel sheet
file_path = "C:\\Users\\HP\\Desktop\\SDET\\Parameterization\\Demo.xlsx"
workbook = load_workbook(file_path)
sheet = workbook["Sheet9"]

# Enter Username
username = sheet.cell(row=1, column=1).value
print(username)                                     # Velocity
driver.find_element(By.XPATH, "//input[@id='username']").send_keys(username)

# Enter Password
password = sheet.cell(row=2, column=1).value
print(password)                                       # U1YP1G
driver.find_element(By.XPATH, "//input[@id='password']").send_keys(password)

# Click On login Button
driver.find_element(By.XPATH, "//input[@id='login']").click()

# Wait
time.sleep(2)

# Expected Username as usual to site
expected_username = sheet.cell(row=1, column=3).value
print(expected_username)

# Actual Username
username_element = driver.find_element(By.XPATH, "//input[@id='username_show']")
username_value = username_element.get_attribute("value")
print(username_value)

# Split the value
List = username_value.split(" ")
a1 = List[0]
print(a1)                   # Hello
actual_username = List[1]
print(actual_username)      # Velocity!

# Check if the usernames match
if expected_username == actual_username:
    print("Test Case Pass!!")
else:
    print("Test Case Fail!!")

# Close the browser
driver.quit()
