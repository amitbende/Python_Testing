import openpyxl


def main():
    file_path = "C:\\Users\\HP\\Desktop\\SDET\\Parameterization\\Demo.xlsx"

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet7"]

    last_cell_index = sheet.max_column - 1
    print(last_cell_index)

    for i in range(0, last_cell_index + 1):
        row_data = sheet.cell(row=1, column=i + 1).value
        print(row_data, end=" ")


if __name__ == "__main__":
    main()
