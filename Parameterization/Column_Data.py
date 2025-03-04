import openpyxl


def main():
    file_path = "C:\\Users\\HP\\Desktop\\SDET\\Parameterization\\Demo.xlsx"

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet7"]

    last_row_index = sheet.max_row
    print(last_row_index)

    for i in range(1, last_row_index + 1):
        column_data = sheet.cell(row=i, column=1).value
        print(column_data)


if __name__ == "__main__":
    main()
