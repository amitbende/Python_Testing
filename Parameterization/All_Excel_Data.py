import openpyxl


def main():
    file_path = "C:\\Users\\HP\\Desktop\\SDET\\Parameterization\\Demo.xlsx"

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet8"]

    last_row_index = sheet.max_row
    last_column_index = sheet.max_column - 1

    for i in range(1, last_row_index + 1):
        row_values = []
        for j in range(0, last_column_index + 1):
            excel_data = sheet.cell(row=i, column=j + 1).value
            row_values.append(excel_data)
        print(" ".join(str(value) for value in row_values))


if __name__ == "__main__":
    main()
